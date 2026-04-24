"""
# v1.1
run_job_pipeline_final.py
=========================
Job alert extraction pipeline — Gmail → Excel

SHEET ARCHITECTURE
------------------
  Jobs    — rolling 3-day window (today and 2 prior calendar days)
  Archive — jobs aged off Jobs sheet; grows permanently; never deleted
  Pivot   — source × date count matrix from Jobs sheet; columns you add
            to the RIGHT of the data region are preserved forever
  <any other sheets you create> — never touched

DEDUPLICATION
-------------
  Checks job_link_clean against Jobs + Archive combined.
  A job that aged into Archive will NOT re-appear as new.

FLOW EACH RUN
-------------
  1. Age-off: rows on Jobs older than today-2 days → appended to Archive
  2. Fetch unprocessed Gmail messages
  3. Parse emails → normalise links → deduplicate
  4. Append new jobs to Jobs sheet
  5. Rebuild Pivot from Jobs sheet (preserving user-added columns)
  6. Apply Job_Processed Gmail label

USAGE
-----
  python run_job_pipeline_final.py --mode historical --dry-run
  python run_job_pipeline_final.py --mode historical
  python run_job_pipeline_final.py --mode forward
"""

import os
import re
import argparse
from datetime import datetime, timezone, timedelta

import pandas as pd
from bs4 import BeautifulSoup
from openpyxl import load_workbook, Workbook
from openpyxl.styles import Font, PatternFill, Alignment

from Email import (
    get_html_body,
    detect_source,
    parse_jooble,
    parse_indeed,
    parse_efinancialcareers,
    parse_adzuna,
)

from gmail_connector import (
    get_gmail_service,
    get_all_label_ids,
    fetch_unprocessed_messages,
    mark_batch_as_processed,
    PROCESSED_LABEL,
)


# ==============================
# CONFIG
# ==============================

MASTER_FILE  = "job_master.xlsx"
LOG_FILE     = "pipeline.log"

SHEET_JOBS    = "Jobs"
SHEET_ARCHIVE = "Archive"
SHEET_PIVOT   = "Pivot"

COLUMNS = ["date", "email_sender", "source", "company", "job_title",
           "job_link", "job_link_clean"]

WINDOW_DAYS = 3   # today + 2 prior calendar days


# ==============================
# LINKEDIN PARSER
# ==============================

def parse_linkedin(soup):
    """
    Handles all three LinkedIn email formats:
      - Standard alert    (jobalerts-noreply) — job alert digest
      - Facet suggestions (jobs-noreply)      — "Explore new jobs for..."
      - Similar jobs      (jobs-noreply)      — "New jobs similar to..."

    Strategy: find every unique jobs/view/{ID} anchor, take find_parent('table')
    for card context, parse title and company from the card text.
    Canonical link: https://www.linkedin.com/jobs/view/{ID}
    """
    jobs = []
    seen_ids = set()

    for a in soup.find_all("a", href=True):
        href = a["href"]
        m = re.search(r"linkedin\.com/comm/jobs/view/(\d+)", href)
        if not m:
            continue

        job_id = m.group(1)
        if job_id in seen_ids:
            continue
        seen_ids.add(job_id)

        # Card context lives in the nearest ancestor <table>
        container = a.find_parent("table")
        if not container:
            continue

        card_text = re.sub(r"\s+", " ", container.get_text(separator=" | ", strip=True))

        # Card format: "Job Title | Company · Location | Salary | ..."
        # Split on " | " then on " · " to separate title from company
        parts = [p.strip() for p in card_text.split(" | ") if p.strip()]
        if not parts:
            continue

        job_title = parts[0]

        # Company is the part before " · " in the second segment
        company = ""
        if len(parts) > 1:
            company_segment = parts[1]
            company = company_segment.split(" · ")[0].strip()

        # Skip navigation/header cards that aren't real jobs
        if not company or job_title.lower().startswith("jobs similar to"):
            continue

        canonical = f"https://www.linkedin.com/jobs/view/{job_id}"

        jobs.append({
            "job_title": job_title,
            "company":   company,
            "job_link":  canonical,
        })

    return jobs


# ==============================
# NORMALISE JOB LINK
# ==============================

def normalise_link(href, source=""):
    """
    Produce a stable, tracking-free canonical URL for deduplication.

    LinkedIn : already canonical from parse_linkedin() — strip query string only
    Indeed   : extract jk= job key → https://uk.indeed.com/viewjob?jk={key}
    Jooble   : strip query string
    Adzuna   : strip query string
    eFC      : strip query string (slug appended separately in parse_msg)
    """
    if not href:
        return href

    if "linkedin.com" in href:
        return href.split("?")[0].split("#")[0].rstrip("/")

    if "indeed.com" in href:
        m = re.search(r"jk=([a-f0-9]+)", href)
        if m:
            return f"https://uk.indeed.com/viewjob?jk={m.group(1)}"

    return href.split("?")[0].split("#")[0].rstrip("/")


# ==============================
# PARSE A SINGLE EMAIL
# ==============================

def parse_msg(msg):
    """
    Parse one email.message object → list of job dicts.
    LinkedIn is handled entirely by parse_linkedin() above.
    All other sources use their existing parsers from Email.py.
    """
    sender  = msg.get("From", "")
    date    = msg.get("Date", "")
    subject = msg.get("Subject", "")
    html    = get_html_body(msg)

    if not html:
        return []

    soup   = BeautifulSoup(html, "lxml")
    source = detect_source(sender)

    if source == "linkedin":
        jobs = parse_linkedin(soup)

    elif source == "jooble":
        jobs = parse_jooble(soup, subject=subject)

    elif source == "indeed":
        jobs = parse_indeed(soup)

    elif source == "efinancialcareers":
        jobs = parse_efinancialcareers(soup)

    elif source == "adzuna":
        jobs = parse_adzuna(soup)

    else:
        return []

    for job in jobs:
        job["date"]         = date
        job["email_sender"] = sender
        job["source"]       = source
        clean = normalise_link(job.get("job_link", ""), source)

        # ── eFC PROTECTION ──────────────────────────────────────────────────
        # eFC uses ONE identical tracking URL for every job in the email.
        # Without this fix, every job gets the same job_link_clean and only
        # the first survives deduplication. Append a title slug to make each
        # key unique. DO NOT REMOVE THIS BLOCK.
        if source == "efinancialcareers" and clean:
            title_slug = re.sub(
                r"[^a-z0-9]+", "-", job.get("job_title", "").lower()
            ).strip("-")
            clean = f"{clean}#{title_slug}"
        # ────────────────────────────────────────────────────────────────────

        job["job_link_clean"] = clean

    return jobs


# ==============================
# EXCEL HELPERS
# ==============================

def _window_start():
    """Oldest date still shown on the Jobs sheet (today - WINDOW_DAYS + 1)."""
    return (datetime.now(timezone.utc) - timedelta(days=WINDOW_DAYS - 1)).date()


def _parse_date_col(series):
    """Robustly parse a date column that may be strings or datetimes."""
    return pd.to_datetime(series, errors="coerce", utc=True).dt.date


def load_workbook_safe(path):
    """Load workbook if it exists, else return a fresh one with no sheets."""
    if os.path.exists(path):
        return load_workbook(path)
    wb = Workbook()
    wb.remove(wb.active)   # remove default Sheet
    return wb


def ensure_sheet(wb, name):
    """Return named sheet, creating it if absent."""
    if name not in wb.sheetnames:
        wb.create_sheet(name)
    return wb[name]


def df_from_sheet(wb, name):
    """Read a sheet into a DataFrame. Returns empty DF if sheet missing/empty."""
    if name not in wb.sheetnames:
        return pd.DataFrame(columns=COLUMNS)
    ws = wb[name]
    data = list(ws.values)
    if not data:
        return pd.DataFrame(columns=COLUMNS)
    headers = list(data[0])
    rows    = data[1:]
    df = pd.DataFrame(rows, columns=headers)
    for col in COLUMNS:
        if col not in df.columns:
            df[col] = ""
    return df[COLUMNS]


def write_df_to_sheet(ws, df):
    """
    Overwrite sheet content with df.
    Clears ALL existing rows, writes header + data.
    Does NOT touch other sheets.
    """
    ws.delete_rows(1, ws.max_row + 1)

    # Header row — bold
    for col_idx, col_name in enumerate(df.columns, start=1):
        cell = ws.cell(row=1, column=col_idx, value=col_name)
        cell.font = Font(bold=True)

    # Data rows
    for row_idx, row in enumerate(df.itertuples(index=False), start=2):
        for col_idx, value in enumerate(row, start=1):
            ws.cell(row=row_idx, column=col_idx, value=value)


def load_all_links(wb):
    """
    Return set of all job_link_clean values across Jobs + Archive.
    Used for deduplication — a job in either sheet is considered seen.
    """
    links = set()
    for sheet_name in (SHEET_JOBS, SHEET_ARCHIVE):
        df = df_from_sheet(wb, sheet_name)
        if "job_link_clean" in df.columns:
            links.update(df["job_link_clean"].dropna().tolist())
    return links


# ==============================
# AGE-OFF: Jobs → Archive
# ==============================

def age_off_jobs(wb, dry_run=False):
    """
    Move rows older than window_start from Jobs sheet → Archive sheet.
    Returns (aged_df, remaining_df).
    """
    cutoff = _window_start()
    df_jobs = df_from_sheet(wb, SHEET_JOBS)

    if df_jobs.empty:
        return pd.DataFrame(columns=COLUMNS), df_jobs

    df_jobs["_date_parsed"] = _parse_date_col(df_jobs["date"])
    aged      = df_jobs[df_jobs["_date_parsed"] < cutoff].drop(columns=["_date_parsed"])
    remaining = df_jobs[df_jobs["_date_parsed"] >= cutoff].drop(columns=["_date_parsed"])

    if aged.empty:
        return aged, remaining

    if dry_run:
        log(f"[DRY RUN] Would age off {len(aged)} rows to Archive")
        return aged, remaining

    # Append aged rows to Archive
    df_archive = df_from_sheet(wb, SHEET_ARCHIVE)
    df_archive_updated = pd.concat([df_archive, aged], ignore_index=True)

    ws_archive = ensure_sheet(wb, SHEET_ARCHIVE)
    write_df_to_sheet(ws_archive, df_archive_updated)

    log(f"Aged off {len(aged)} rows to {SHEET_ARCHIVE}")
    return aged, remaining


# ==============================
# DEDUPLICATE + APPEND
# ==============================

def append_new_jobs(wb, new_jobs, existing_links, remaining_jobs_df, dry_run=False):
    """
    Filter new_jobs against existing_links, append fresh ones to Jobs sheet.
    Returns (new_df, updated_existing_links).
    """
    fresh = [j for j in new_jobs if j.get("job_link_clean") not in existing_links]

    if not fresh:
        return pd.DataFrame(), existing_links

    df_new = pd.DataFrame(fresh)
    for col in COLUMNS:
        if col not in df_new.columns:
            df_new[col] = ""
    df_new = df_new[COLUMNS]

    # Parse and sort dates
    df_new["date"] = pd.to_datetime(df_new["date"], errors="coerce", utc=True)
    df_new = df_new.sort_values(["date", "company"], ascending=[False, True])
    df_new["date"] = df_new["date"].dt.strftime("%Y-%m-%d")

    if dry_run:
        log(f"[DRY RUN] Would append {len(df_new)} new jobs to {SHEET_JOBS}")
        log("[DRY RUN] Preview:")
        print(df_new[["date", "source", "company", "job_title"]].to_string(index=False))
        updated_links = existing_links | set(df_new["job_link_clean"].dropna())
        return df_new, updated_links

    # Combine remaining jobs + new jobs, write back to Jobs sheet
    df_combined = pd.concat([remaining_jobs_df, df_new], ignore_index=True)
    df_combined["_date_parsed"] = _parse_date_col(df_combined["date"])
    df_combined = df_combined.sort_values(["_date_parsed", "company"], ascending=[False, True])
    df_combined = df_combined.drop(columns=["_date_parsed"])

    ws_jobs = ensure_sheet(wb, SHEET_JOBS)
    write_df_to_sheet(ws_jobs, df_combined)

    updated_links = existing_links | set(df_new["job_link_clean"].dropna())
    return df_new, updated_links


# ==============================
# REBUILD PIVOT
# ==============================

# Fixed source columns — always present in this order regardless of what's in the data
PIVOT_SOURCES = ["adzuna", "efinancialcareers", "indeed", "jooble", "linkedin"]
PIVOT_FIXED_COLS = ["company", "job_title"] + PIVOT_SOURCES   # 7 columns we own


def rebuild_pivot(wb, dry_run=False):
    """
    Rebuild the Pivot sheet from the current Jobs sheet.

    Layout:
        Company | Job Title | adzuna | efinancialcareers | indeed | jooble | linkedin
        One row per unique Company + Job Title combination.
        URL cell = job_link for that source; blank if that source didn't have the job.
        Same job on multiple sources → one row, multiple URLs filled.

    User-added columns to the RIGHT of column 7 (linkedin) are preserved forever.
    Jobs and Archive sheets are never touched.
    """
    if dry_run:
        log("[DRY RUN] Would rebuild Pivot sheet")
        return

    df_jobs = df_from_sheet(wb, SHEET_JOBS)

    # ── PRESERVE USER COLUMNS ────────────────────────────────────────────────
    # Save anything in columns 8+ before we clear the sheet.
    user_extra_cols = {}   # {col_idx: [val_row1, val_row2, ...]}

    if SHEET_PIVOT in wb.sheetnames:
        ws_existing = wb[SHEET_PIVOT]
        max_col = ws_existing.max_column or 0
        if max_col > len(PIVOT_FIXED_COLS):
            for col_idx in range(len(PIVOT_FIXED_COLS) + 1, max_col + 1):
                col_data = [
                    ws_existing.cell(row=r, column=col_idx).value
                    for r in range(1, ws_existing.max_row + 1)
                ]
                user_extra_cols[col_idx] = col_data
    # ─────────────────────────────────────────────────────────────────────────

    ws_pivot = ensure_sheet(wb, SHEET_PIVOT)
    ws_pivot.delete_rows(1, ws_pivot.max_row + 1)

    if df_jobs.empty:
        ws_pivot.cell(row=1, column=1, value="No data in Jobs sheet")
        log("Pivot sheet rebuilt (empty — no jobs in window)")
        return

    # Build the pivot: one row per company+job_title, one column per source
    # Use job_link (full URL) as the cell value
    pivot_data = {}   # {(company, job_title): {source: url}}

    for _, row in df_jobs.iterrows():
        company   = str(row.get("company", "") or "").strip()
        job_title = str(row.get("job_title", "") or "").strip()
        source    = str(row.get("source", "") or "").strip().lower()
        job_link  = str(row.get("job_link", "") or "").strip()

        if not company or not job_title:
            continue

        key = (company, job_title)
        if key not in pivot_data:
            pivot_data[key] = {}

        # Only fill if source is one we track and cell not already populated
        if source in PIVOT_SOURCES and source not in pivot_data[key]:
            pivot_data[key][source] = job_link

    # Sort rows: company then job_title
    sorted_keys = sorted(pivot_data.keys(), key=lambda x: (x[0].lower(), x[1].lower()))

    # Write header row
    header_fill = PatternFill("solid", start_color="1F4E79")
    header_font = Font(bold=True, color="FFFFFF")
    headers = ["Company", "Job Title"] + PIVOT_SOURCES

    for col_idx, col_name in enumerate(headers, start=1):
        cell = ws_pivot.cell(row=1, column=col_idx, value=col_name)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center")

    # Set sensible column widths
    ws_pivot.column_dimensions["A"].width = 35   # Company
    ws_pivot.column_dimensions["B"].width = 45   # Job Title
    for col_letter in ["C", "D", "E", "F", "G"]:
        ws_pivot.column_dimensions[col_letter].width = 20

    # Write data rows
    for row_idx, key in enumerate(sorted_keys, start=2):
        company, job_title = key
        source_urls = pivot_data[key]

        ws_pivot.cell(row=row_idx, column=1, value=company)
        ws_pivot.cell(row=row_idx, column=2, value=job_title)

        for col_idx, source in enumerate(PIVOT_SOURCES, start=3):
            url = source_urls.get(source, "")
            cell = ws_pivot.cell(row=row_idx, column=col_idx, value=url)
            # Make URLs wrap rather than overflow
            if url:
                cell.alignment = Alignment(wrap_text=True)

    # Re-write user columns to the right of our data region
    for col_idx, col_data in user_extra_cols.items():
        for row_idx, value in enumerate(col_data, start=1):
            ws_pivot.cell(row=row_idx, column=col_idx, value=value)

    log(f"Pivot sheet rebuilt ({len(sorted_keys)} unique company+title rows)")


# ==============================
# LOGGING
# ==============================

def log(message):
    ts   = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
    line = f"[{ts}] {message}"
    print(line)
    with open(LOG_FILE, "a", encoding="utf-8") as f:
        f.write(line + "\n")


# ==============================
# MAIN PIPELINE
# ==============================

def run_pipeline(mode="forward", dry_run=False):
    log("=" * 60)
    log(f"Pipeline start | mode={mode} | dry_run={dry_run}")
    log(f"Window: {_window_start()} → today")

    # 1. Connect to Gmail
    service   = get_gmail_service()
    label_ids = get_all_label_ids(service)

    # 2. Load workbook (creates fresh if missing)
    wb = load_workbook_safe(MASTER_FILE)

    # 3. Age off old rows from Jobs → Archive
    log("\nChecking for rows to age off...")
    aged_df, remaining_df = age_off_jobs(wb, dry_run=dry_run)
    if not dry_run and not aged_df.empty:
        ws_jobs = ensure_sheet(wb, SHEET_JOBS)
        write_df_to_sheet(ws_jobs, remaining_df)
        log(f"Removed {len(aged_df)} aged rows from {SHEET_JOBS}")

    # 4. Build combined dedup set from Jobs + Archive
    existing_links = load_all_links(wb)
    log(f"Dedup pool: {len(existing_links)} links (Jobs + Archive)")

    # 5. Fetch unprocessed Gmail messages
    log("\nFetching unprocessed messages from Gmail...")
    messages = fetch_unprocessed_messages(service, label_ids, dry_run=dry_run)
    log(f"Messages to process: {len(messages)}")

    if not messages:
        log("No new messages.")
    else:
        # 6. Parse
        log("\nParsing emails...")
        all_new_jobs      = []
        processed_msg_ids = []

        for msg_id, msg_obj, label_name in messages:
            jobs = parse_msg(msg_obj)
            processed_msg_ids.append(msg_id)   # mark processed regardless

            if jobs:
                all_new_jobs.extend(jobs)
                print(f"  ✓ {label_name} | {len(jobs)} jobs | {msg_id}")
            else:
                print(f"  - {label_name} | 0 jobs  | {msg_id}")

        log(f"Total jobs parsed: {len(all_new_jobs)}")

        # 7. Deduplicate + append
        df_new, existing_links = append_new_jobs(
            wb, all_new_jobs, existing_links, remaining_df, dry_run=dry_run
        )

        if df_new.empty:
            log("No new jobs after deduplication.")
        else:
            log(f"✓ {len(df_new)} new jobs added to {SHEET_JOBS}")

        # 8. Label Gmail messages
        if processed_msg_ids:
            log(f"\nLabelling {len(processed_msg_ids)} messages as '{PROCESSED_LABEL}'...")
            mark_batch_as_processed(service, processed_msg_ids, label_ids, dry_run=dry_run)

    # 9. Rebuild pivot
    log("\nRebuilding Pivot sheet...")
    rebuild_pivot(wb, dry_run=dry_run)

    # 10. Save workbook
    if not dry_run:
        wb.save(MASTER_FILE)
        log(f"✓ Saved {MASTER_FILE}")

    log(f"\nPipeline complete")
    log("=" * 60)


# ==============================
# CLI
# ==============================

if __name__ == "__main__":
    parser = argparse.ArgumentParser(description="Job alert email pipeline")
    parser.add_argument(
        "--mode", choices=["historical", "forward"], default="forward",
        help="historical: bulk-process all unprocessed emails. forward: run on schedule."
    )
    parser.add_argument(
        "--dry-run", action="store_true",
        help="Preview without writing to Excel or labelling Gmail messages."
    )
    args = parser.parse_args()
    run_pipeline(mode=args.mode, dry_run=args.dry_run)
